import calendar
import io
import os
from datetime import date

import requests
from django.utils.translation import gettext as _
from docx.shared import Mm
from docxtpl import InlineImage, DocxTemplate
from openpyxl import Workbook

from BackStrawBerryPy.models import BaseModel
from BackStrawBerryPy.settings import BASE_DIR
from apps.Matriculas.models import AlumnoAula
from apps.Notas.models import NotaAlumno, EvidenciaNotaAlumno
from apps.Personas.models import Personal
from utils.functions import create_docx, docx_to_bytes, create_docx_bytes, get_edad, concat_if_exist, get_ordinal_num

DIR_REPORTES = os.path.join(BASE_DIR, 'static', 'reportes')


def get_reporte(reporte):
    return os.path.join(DIR_REPORTES, reporte)


def reporte_nomina():
    personal = Personal.objects.filter(auth_estado=BaseModel.ACTIVO)

    docx = create_docx(
        file=get_reporte('MatrizPersonalIpca.docx'),
        context={
            'personal': [
                dict(
                    index=i + 1,
                    nombres=item.persona.get_nombres_apellidos(),
                    identificacion=item.persona.identificacion,
                    titulo_profecional=item.titulo,
                    celular=item.persona.celular_uno,
                    correo=item.persona.correo,
                    funcion=item.get_funciones_str()
                )
                for i, item in enumerate(personal)
            ]
        }
    )
    file_stream = io.BytesIO()
    docx.save(file_stream)
    file_stream.seek(0)
    return file_stream


def reporte_notas(id_matricula, id_docente, area_trabajo, responsable):
    matricula = AlumnoAula.objects.get(pk=id_matricula)

    notas = NotaAlumno.objects.prefetch_related("evidencias").filter(
        alumno_aula_id=id_matricula,
        auth_estado=BaseModel.ACTIVO
    ).order_by('componente__nombre')

    imgs = EvidenciaNotaAlumno.objects.filter(
        nota_id__in=[nota.pk for nota in notas]
    ).values_list('url', flat=True)

    tpl = DocxTemplate(get_reporte('MATRIZ_INFORME.docx'))
    imgs_objs = dict()

    for index, img in enumerate(list(imgs)):
        response = requests.get(img, stream=True)
        image = io.BytesIO(response.content)
        my_image = InlineImage(tpl, image, width=Mm(150))
        imgs_objs.__setitem__(f'img{index + 1}', my_image)

    tpl.render({**{
        'responsable': responsable,
        'mes': _(calendar.month_name[date.today().month]).upper(),
        'area_de_trabajo': area_trabajo,
        'docente': Personal.objects.select_related('persona').get(pk=id_docente).persona.__str__(),
        'coordinador': matricula.aula.periodo.coordinador.persona.__str__(),
        'director': matricula.aula.periodo.director.persona.__str__(),
        'notas': [
            dict(
                index=i + 1,
                componente=item.componente.nombre,
                actividad=item.titulo,
                resultado=item.resultado,
                observacion=item.observaciones,
            )
            for i, item in enumerate(notas)
        ],
    }, **imgs_objs})
    return docx_to_bytes(tpl)


def reporte_ficha_inscripcion(id_matricula):
    matricula = AlumnoAula.objects \
        .select_related('alumno__persona', 'aula') \
        .get(pk=id_matricula)

    persona = matricula.alumno.persona
    aula = matricula.aula
    alumno = matricula.alumno

    return create_docx_bytes(
        file=get_reporte("FichaInscripcion.docx"),
        context={
            'anio': date.today().year,
            'apellidos': persona.get_apellidos(),
            'nombres': persona.get_nombres(),
            'lugar': persona.direccion_domiciliaria,
            'fecha_n': persona.fecha_nacimiento.strftime('%d/%m/%Y'),
            'edad': get_edad(persona.fecha_nacimiento),
            'cedula': persona.identificacion,
            'conadis': persona.carnet_conadis,
            'nivel_a': _(get_ordinal_num(aula.grado)),
            'promovido': _(get_ordinal_num(aula.grado)),
            'tratamiento': matricula.tratamiento,
            'diagnostico': matricula.diagnostico_clinico,
            # INFORMACION DEL PADRE
            'apellidos_p': concat_if_exist(alumno.padre.get('primer_apellido'), alumno.padre.get('segundo_apellido')),
            'nombres_p': concat_if_exist(alumno.padre.get('primer_nombre'), alumno.padre.get('segundo_nombre')),
            'cedula_p': alumno.padre.get('identificacion'),
            'ocupacion_p': alumno.padre.get('ocupacion'),
            'direccion_p': alumno.padre.get('direccion'),
            'telefono_p': alumno.padre.get('telefono'),
            'celular_p': alumno.padre.get('celular'),
            # INFORMACION DE LA MADRE
            'apellidos_m': concat_if_exist(alumno.madre.get('primer_apellido'), alumno.madre.get('segundo_apellido')),
            'nombres_m': concat_if_exist(alumno.madre.get('primer_nombre'), alumno.madre.get('segundo_nombre')),
            'cedula_m': alumno.madre.get('identificacion'),
            'ocupacion_m': alumno.madre.get('ocupacion'),
            'direccion_m': alumno.madre.get('direccion'),
            'telefono_m': alumno.madre.get('telefono'),
            'celular_m': alumno.madre.get('celular'),
            'correo': persona.correo,
            'direccion': persona.direccion_domiciliaria,
            'provincia': persona.provincia_residencia,
            'canton': persona.canton_residencia,
            'parroquia': persona.parroquia_residencia,
            'sector': persona.sector,
            # CONTACTO_EMERGENCIA
            'nombre_emergencia': alumno.contacto_emergencia.get('nombres'),
            'contacto': alumno.contacto_emergencia.get('contacto'),
            'nombre_representante': alumno.representante.get('nombres'),
            # MATRICULA
            'matricula': matricula.numero_matricula,
            'aporte': f"${matricula.aporte_voluntario}",
            'fecha_inscripcion': matricula.created_at.strftime('%d/%m/%Y')
        }
    )


def set_column(array, value):
    array.append(value)


def reporte_general_total_alumnos(id_periodo):
    matriculas = AlumnoAula.objects \
        .select_related('alumno', 'alumno__persona', 'aula', 'aula__periodo') \
        .filter(auth_estado=BaseModel.ACTIVO, aula__periodo__id__in=[id_periodo])

    rows = []

    for index, matricula in enumerate(matriculas):
        alumno = matricula.alumno
        persona = alumno.persona
        aula = matricula.aula
        array = []
        set_column(array, index + 1)
        set_column(array, persona.get_nombres_apellidos())
        set_column(array, matricula.amie)
        set_column(array, matricula.mies)
        set_column(array, aula.nombre)
        set_column(array, alumno.created_at.strftime('%d/%m/%Y'))
        set_column(array, f"{persona.pais_nacimiento}")
        set_column(array, persona.fecha_nacimiento.strftime('%d/%m/%Y'))
        set_column(array, persona.get_edad())
        set_column(array, persona.identificacion)
        set_column(array, persona.genero)
        set_column(array, persona.estado_civil)
        set_column(array, persona.tiene_discapacidad)
        set_column(array, persona.carnet_conadis)
        set_column(array, "")  # TODO: mapp discapacidades
        set_column(array, persona.porcentaje_discapacidad)
        set_column(array, persona.etnia)
        set_column(array, alumno.historia_clinica)
        set_column(array, persona.tipo_sangre)
        set_column(array, matricula.diagnostico_clinico)
        set_column(array, alumno.trastornos_asociados)
        set_column(array, matricula.grado_dependencia)
        set_column(array, alumno.representante.get("parentesco", ""))
        set_column(array, alumno.map_padres().get('apellidos_nombres'))
        set_column(array, alumno.padre.get('identificacion'))
        set_column(array, alumno.map_padres("madre").get("apellidos_nombres"))
        set_column(array, alumno.madre.get("identificacion"))
        set_column(array, matricula.tipo_familia)
        set_column(array, alumno.bono)
        set_column(array, alumno.tipo_bono)
        set_column(array, alumno.afiliacion_iess)
        set_column(array, alumno.quintil_pobreza)
        set_column(array, persona.pais_residencia)
        set_column(array, persona.provincia_residencia)
        set_column(array, persona.canton_residencia)
        set_column(array, persona.parroquia_residencia)
        set_column(array, persona.direccion_domiciliaria)
        set_column(array, persona.telefono)
        set_column(array, persona.celular_uno)
        set_column(array, persona.celular_dos)

        rows.append(array)

    columns = [
        "No",
        "NOMBRE DEL USUARIO",
        "AMIE",
        "MIES",
        "NIVELES",
        "FECHA DE INGRESO",
        "PAIS DE NACIMIENTO",
        "FECHA DE NACIMIENTO USUARIO",
        "EDAD",
        # "MESES",
        "CEDULA O PASAPORTE",
        "GENERO",
        "ESTADO CIVIL",
        "CARNET DE DISCAPACIDAD",
        "Nª DE REGISTRO CARNET",
        "TIPO DE DISCAPACIDAD",
        "PORCENTAJE",
        "ETNIA",
        "HISTORIA CLINICA",
        "TIPO DE SANGRE",
        "DIAGNOSTICO CLINICO",
        "TRASTORNOS ASOCIADOS",
        "GRADO DE DEPENDENCIA",
        "REPRESENTANTE",
        "NOMBRE PADRE",
        "CI PADRE",
        "NOMBRE MADRE",
        "C.I MADRE",
        "TIPO DE FLIA",
        "BONO",
        "TIPO DE BONO",
        "AFILICACIÓN IESS",
        "QUINTIL DE POBREZA",
        "PAIS",
        "PROVINCIA",
        "CANTON",
        "PARROQUIA",
        "DIRECCION",
        "TELEFONO DOMICILIO",
        "CELULAR 1",
        "CELULAR 2",
    ]

    workbook = Workbook()
    worksheet = workbook.active

    worksheet.title = "Repore general de Alumnos"
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for row in rows:
        row_num += 1
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    buffer = io.BytesIO()

    workbook.save(buffer)

    buffer.seek(0)

    return buffer
